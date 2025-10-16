#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Bilanço Uygulaması (Python CLI + Tkinter GUI)
- TSX bileşenindeki hesaplamalar ve doğrulamalar birebir korunmuştur.
- CLI: Kullanıcıdan verileri alır, hesaplar ve doğrular.
- GUI: Girdi alanları, canlı hesaplama ve Canvas ile görselleştirme; doğrulama ve JSON kaydetme.
"""

from __future__ import annotations

import json
import sys
import argparse
from datetime import datetime
from typing import Dict, List, Tuple
import os
import unicodedata
import difflib
try:
  from pymongo import MongoClient  # type: ignore
except Exception:
  MongoClient = None  # type: ignore

try:
  import tkinter as tk
  from tkinter import ttk, messagebox, filedialog
  try:
    from tkinter import simpledialog  # type: ignore
  except Exception:
    simpledialog = None  # type: ignore
except Exception:
  tk = None
  ttk = None
  messagebox = None
  filedialog = None
  simpledialog = None  # type: ignore

# Optional Excel support
try:
  from openpyxl import Workbook, load_workbook  # type: ignore
except Exception:
  Workbook = None  # type: ignore
  load_workbook = None  # type: ignore


# Kullanıcı girdilerini güvenle sayıya çevirme
def parse_float(value: str) -> float:
  try:
    normalized = value.strip().replace(" ", "").replace(",", ".")
    if normalized == "":
      return 0.0
    return float(normalized)
  except Exception:
    return 0.0


def prompt_str(prompt: str, default: str = "") -> str:
  raw = input(f"{prompt} [{default}]: ")
  return raw.strip() if raw.strip() != "" else default


def prompt_num(prompt: str, default: float = 0.0) -> float:
  raw = input(f"{prompt} [{default}]: ")
  return parse_float(raw) if raw.strip() != "" else float(default)


def collect_bilanco_data() -> Dict[str, float | str]:
  today = datetime.now().strftime("%Y-%m-%d")
  print("\nBilgi girişine başlayalım. Boş bırakılan alanlar varsayılan 0 kabul edilir.\n")

  data: Dict[str, float | str] = {
    "isletmeAdi": prompt_str("İşletme Adı / Ünvanı", ""),
    "bilancoTarihi": prompt_str("Bilanço Tarihi (YYYY-MM-DD)", today),

    # I. Dönen Varlıklar
    "kasa": prompt_num("100 - Kasa", 0.0),
    "bankalar": prompt_num("102 - Bankalar", 0.0),
    "alicilar": prompt_num("120 - Alıcılar", 0.0),
    "alacakSenetleri": prompt_num("121 - Alacak Senetleri", 0.0),
    "verilenDepozito": prompt_num("126 - Verilen Depozito ve Teminatlar", 0.0),
    "digerAlacaklar": prompt_num("136 - Diğer Çeşitli Alacaklar", 0.0),
    "ticariMallar": prompt_num("153 - Ticari Mallar", 0.0),
    "yariMamul": prompt_num("154 - Yarı Mamuller", 0.0),
    "mamul": prompt_num("155 - Mamuller", 0.0),
    "digerDonenVarliklar": prompt_num("199 - Diğer Dönen Varlıklar", 0.0),

    # II. Duran Varlıklar
    "ticariAlacaklar": prompt_num("220 - Ticari Alacaklar (Uzun Vadeli)", 0.0),
    "istirakler": prompt_num("242 - İştirakler", 0.0),
    "bagliOrtakliklar": prompt_num("245 - Bağlı Ortaklıklar", 0.0),
    "arazi": prompt_num("250 - Arazi ve Arsalar", 0.0),
    "binalar": prompt_num("252 - Binalar", 0.0),
    "tesisatMakineler": prompt_num("253 - Tesis, Makine ve Cihazlar", 0.0),
    "demirbaslar": prompt_num("255 - Demirbaşlar", 0.0),
    "tasitlar": prompt_num("254 - Taşıtlar", 0.0),
    "birikmiAmort": prompt_num("257 - Birikmiş Amortismanlar (-)", 0.0),
    "digerDuranVarliklar": prompt_num("299 - Diğer Duran Varlıklar", 0.0),

    # III. Kısa Vadeli Yabancı Kaynaklar
    "bankKredileri": prompt_num("300 - Banka Kredileri", 0.0),
    "saticilar": prompt_num("320 - Satıcılar", 0.0),
    "borcSenetleri": prompt_num("321 - Borç Senetleri", 0.0),
    "digerBorclar": prompt_num("336 - Diğer Çeşitli Borçlar", 0.0),
    "odenecekVergiler": prompt_num("360 - Ödenecek Vergi ve Fonlar", 0.0),

    # IV. Uzun Vadeli Yabancı Kaynaklar
    "uzunVadeBankKredileri": prompt_num("400 - Banka Kredileri (Uzun Vadeli)", 0.0),
    "tahviller": prompt_num("420 - Çıkarılmış Tahviller", 0.0),
    "uzunVadeBorclar": prompt_num("436 - Diğer Borçlar", 0.0),

    # V. Öz Kaynaklar
    "odenmisSermaye": prompt_num("500 - Ödenmiş Sermaye", 0.0),
    "sermayeYedekleri": prompt_num("520 - Sermaye Yedekleri", 0.0),
    "karYedekleri": prompt_num("540 - Kar Yedekleri", 0.0),
    "gecmisYilKarlari": prompt_num("570 - Geçmiş Yıl Karları", 0.0),
    "donemNetKari": prompt_num("590 - Dönem Net Karı", 0.0),
  }

  return data


def sum_donen_varliklar(d: Dict[str, float | str]) -> float:
  total = 0.0
  for key in DONEN_KEYS:
    try:
      total += float(d.get(key, 0) or 0)
    except Exception:
      continue
  return total


def sum_duran_varliklar(d: Dict[str, float | str]) -> float:
  total = 0.0
  for key in DURAN_KEYS:
    try:
      total += float(d.get(key, 0) or 0)
    except Exception:
      continue
  try:
    total -= float(d.get("birikmiAmort", 0) or 0)
  except Exception:
    pass
  return total


def sum_kv_yabanci_kaynaklar(d: Dict[str, float | str]) -> float:
  total = 0.0
  for key in KV_KEYS:
    try:
      total += float(d.get(key, 0) or 0)
    except Exception:
      continue
  return total


def sum_uv_yabanci_kaynaklar(d: Dict[str, float | str]) -> float:
  total = 0.0
  for key in UV_KEYS:
    try:
      total += float(d.get(key, 0) or 0)
    except Exception:
      continue
  return total


def sum_oz_kaynaklar(d: Dict[str, float | str]) -> float:
  total = 0.0
  for key in OZ_KEYS:
    try:
      total += float(d.get(key, 0) or 0)
    except Exception:
      continue
  return total


def format_tl(value: float) -> str:
  return f"{value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def validate(d: Dict[str, float | str]) -> List[Dict[str, str]]:
  errors: List[Dict[str, str]] = []

  aktif = sum_donen_varliklar(d) + sum_duran_varliklar(d)
  pasif = sum_kv_yabanci_kaynaklar(d) + sum_uv_yabanci_kaynaklar(d) + sum_oz_kaynaklar(d)

  if abs(aktif - pasif) > 0.01:
    errors.append({
      "tip": "kritik",
      "mesaj": f"Bilanço dengesizliği: Aktif ({format_tl(aktif)} TL) != Pasif ({format_tl(pasif)} TL). Fark: {format_tl(aktif - pasif)} TL"
    })

  if not str(d.get("isletmeAdi", "")).strip():
    errors.append({"tip": "uyari", "mesaj": "İşletme adı boş bırakılmamalıdır."})

  for key, value in d.items():
    if key in ("isletmeAdi", "bilancoTarihi", "birikmiAmort"):
      continue
    try:
      num = float(value)
    except Exception:
      continue
    if num < 0:
      errors.append({"tip": "uyari", "mesaj": f"{key} negatif değer içeriyor: {num}"})

  kv = sum_kv_yabanci_kaynaklar(d)
  dv = sum_donen_varliklar(d)
  likidite_orani = dv / (kv if kv != 0 else 1)
  if likidite_orani < 1.0:
    errors.append({
      "tip": "uyari",
      "mesaj": f"Likidite oranı düşük ({likidite_orani:.2f}). Kısa vadeli borçlar dönen varlıklardan fazla."
    })

  if sum_oz_kaynaklar(d) < 0:
    errors.append({"tip": "kritik", "mesaj": "Öz kaynaklar negatif! İşletme mali sıkıntı içinde olabilir."})

  if aktif == 0:
    errors.append({"tip": "uyari", "mesaj": "Hiçbir varlık girilmemiş. Bilanço boş görünüyor."})

  return errors


def build_mongo_like_document(d: Dict[str, float | str], errors: List[Dict[str, str]]) -> Dict:
  donen = sum_donen_varliklar(d)
  duran = sum_duran_varliklar(d)
  aktif_toplam = donen + duran
  kv = sum_kv_yabanci_kaynaklar(d)
  uv = sum_uv_yabanci_kaynaklar(d)
  oz = sum_oz_kaynaklar(d)
  pasif_toplam = kv + uv + oz

  likidite_orani = donen / (kv if kv != 0 else 1)
  ozkaynaklar_orani = (oz / aktif_toplam * 100) if aktif_toplam != 0 else 0.0
  borc_orani = ((kv + uv) / aktif_toplam * 100) if aktif_toplam != 0 else 0.0

  doc = {
    "_id": None,  # MongoDB kayıt edilirken artan tamsayı atanacak
    "isletmeBilgileri": {
      "ad": d.get("isletmeAdi", ""),
      "tarih": d.get("bilancoTarihi", ""),
    },
    "aktif": {
      "donenVarliklar": {
        "kasa": d.get("kasa", 0),
        "bankalar": d.get("bankalar", 0),
        "alicilar": d.get("alicilar", 0),
        "alacakSenetleri": d.get("alacakSenetleri", 0),
        "verilenDepozito": d.get("verilenDepozito", 0),
        "digerAlacaklar": d.get("digerAlacaklar", 0),
        "ticariMallar": d.get("ticariMallar", 0),
        "yariMamul": d.get("yariMamul", 0),
        "mamul": d.get("mamul", 0),
        "digerDonenVarliklar": d.get("digerDonenVarliklar", 0),
        "toplam": donen,
      },
      "duranVarliklar": {
        "ticariAlacaklar": d.get("ticariAlacaklar", 0),
        "istirakler": d.get("istirakler", 0),
        "bagliOrtakliklar": d.get("bagliOrtakliklar", 0),
        "arazi": d.get("arazi", 0),
        "binalar": d.get("binalar", 0),
        "tesisatMakineler": d.get("tesisatMakineler", 0),
        "demirbaslar": d.get("demirbaslar", 0),
        "tasitlar": d.get("tasitlar", 0),
        "birikmiAmort": d.get("birikmiAmort", 0),
        "digerDuranVarliklar": d.get("digerDuranVarliklar", 0),
        "toplam": duran,
      },
      "toplam": aktif_toplam,
    },
    "pasif": {
      "kisaVadeliYabanciKaynaklar": {
        "bankKredileri": d.get("bankKredileri", 0),
        "saticilar": d.get("saticilar", 0),
        "borcSenetleri": d.get("borcSenetleri", 0),
        "digerBorclar": d.get("digerBorclar", 0),
        "odenecekVergiler": d.get("odenecekVergiler", 0),
        "toplam": kv,
      },
      "uzunVadeliYabanciKaynaklar": {
        "uzunVadeBankKredileri": d.get("uzunVadeBankKredileri", 0),
        "tahviller": d.get("tahviller", 0),
        "uzunVadeBorclar": d.get("uzunVadeBorclar", 0),
        "toplam": uv,
      },
      "ozKaynaklar": {
        "odenmisSermaye": d.get("odenmisSermaye", 0),
        "sermayeYedekleri": d.get("sermayeYedekleri", 0),
        "karYedekleri": d.get("karYedekleri", 0),
        "gecmisYilKarlari": d.get("gecmisYilKarlari", 0),
        "donemNetKari": d.get("donemNetKari", 0),
        "toplam": oz,
      },
      "toplam": pasif_toplam,
    },
    "rasyolar": {
      "likiditeOrani": f"{likidite_orani:.2f}",
      "ozkaynaklarOrani": f"{ozkaynaklar_orani:.2f}",
      "borcOrani": f"{borc_orani:.2f}",
    },
    "kayitTarihi": datetime.now().isoformat(),
    "dogrulama": {
      "durumu": "basarili" if len(errors) == 0 else "uyarilarla",
      "hatalar": errors,
    },
  }
  return doc


def save_to_mongo(doc: Dict, uri: str, db_name: str, coll_name: str) -> str:
  if MongoClient is None:
    raise RuntimeError("PyMongo yüklü değil. 'pip install pymongo' ile yükleyin.")
  client = MongoClient(uri)
  try:
    coll = client[db_name][coll_name]
    # Basit artan tamsayı _id kullan (1,2,3,...) – mevcut en büyük tamsayıya +1
    def _compute_next_id() -> int:
      next_id = 1
      try:
        cursor = coll.find({}, projection={"_id": 1}).sort("_id", -1).limit(200)
        for item in cursor:
          val = item.get("_id")
          if isinstance(val, int):
            next_id = int(val) + 1
            break
      except Exception:
        pass
      return next_id

    payload = dict(doc)
    if not isinstance(payload.get("_id"), int):
      payload["_id"] = _compute_next_id()
    result = coll.insert_one(payload)
    return str(result.inserted_id)
  finally:
    client.close()


def build_excel_rows_from_data(d: Dict[str, float | str]) -> List[Tuple[str, str, str, str, float]]:
  rows: List[Tuple[str, str, str, str, float]] = []
  for side in ("AKTIF", "PASIF"):
    for group, fields in SECTION_FIELDS[side].items():
      for label, key in fields:
        if key.endswith("_dummy"):
          continue
        try:
          val = float(d.get(key, 0) or 0)
        except Exception:
          val = 0.0
        rows.append((side, group, label, key, val))
  return rows


def save_to_excel(path: str, d: Dict[str, float | str]) -> None:
  if Workbook is None:
    raise RuntimeError("openpyxl yüklü değil. 'pip install openpyxl' ile yükleyin.")
  wb = Workbook()
  ws = wb.active
  ws.title = "Bilanço"
  ws.append(["Taraf", "Grup", "Etiket", "Anahtar", "Tutar"])  # header
  for side, group, label, key, val in build_excel_rows_from_data(d):
    ws.append([side, group, label, key, val])
  # meta sheet
  meta = wb.create_sheet("Bilgi")
  meta["A1"] = "İşletme Adı"
  meta["B1"] = str(d.get("isletmeAdi", ""))
  meta["A2"] = "Bilanço Tarihi"
  meta["B2"] = str(d.get("bilancoTarihi", ""))
  wb.save(path)


def load_from_excel(path: str) -> Dict[str, float | str]:
  if load_workbook is None:
    raise RuntimeError("openpyxl yüklü değil. 'pip install openpyxl' ile yükleyin.")
  wb = load_workbook(path, data_only=True)
  d: Dict[str, float | str] = {}
  # meta
  if "Bilgi" in wb.sheetnames:
    meta = wb["Bilgi"]
    d["isletmeAdi"] = str(meta["B1"].value or "")
    d["bilancoTarihi"] = str(meta["B2"].value or "")
  # rows
  ws = wb[wb.sheetnames[0]]
  
  # Detect format: simple 2-column (Label, Value) or detailed 5-column format
  header_row_idx = 1
  is_simple_format = False
  
  # Check if it's simple 2-column format
  if ws.max_column <= 2:
    is_simple_format = True
    label_col = 1
    val_col = 2
  else:
    # Detect detailed format headers
    for r_idx in range(1, min(10, ws.max_row) + 1):
      row_vals = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[r_idx]]
      if any(x for x in row_vals):
        if "anahtar" in row_vals and "tutar" in row_vals:
          header_row_idx = r_idx
          key_col = row_vals.index("anahtar") + 1
          val_col = row_vals.index("tutar") + 1
          if "etiket" in row_vals:
            label_col = row_vals.index("etiket") + 1
          else:
            label_col = 3  # guess
          break
    if not is_simple_format:
      # fallback to default positions for detailed format
      key_col = 4
      val_col = 5
      label_col = 3

  # Normalization helpers for robust label matching
  def _normalize(s: str) -> str:
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))  # remove accents
    allowed = "abcdefghijklmnopqrstuvwxyz0123456789 "
    s = "".join(ch if ch in allowed else " " for ch in s)
    s = " ".join(s.split())
    return s

  # Fuzzy eşleşmeyi kapatmak için katı etiket eşleme modu.
  # Ekrandaki görünen başlıklarla birebir aynı olmayan etiketler eşlenmez.
  STRICT_LABELS = True

  # Build label->key mapping from our SECTION_FIELDS (normalized)
  label_to_key: Dict[str, str] = {}
  # Group-specific mappings to resolve duplicate labels across groups (e.g., "Diğer Borçlar")
  group_label_to_key_map: Dict[str, Dict[str, str]] = {}
  group_names_norm: set[str] = set()
  for side in ("AKTIF", "PASIF"):
    for group, fields in SECTION_FIELDS[side].items():
      group_norm = _normalize(group)
      group_names_norm.add(group_norm)
      if group_norm not in group_label_to_key_map:
        group_label_to_key_map[group_norm] = {}
      for label, key in fields:
        if key.endswith("_dummy"):
          continue
        norm_label = _normalize(label)
        label_to_key[norm_label] = key
        group_label_to_key_map[group_norm][norm_label] = key

  # Common aliases for label variations seen in Excel exports
  # Map alias (normalized) -> canonical key
  alias_to_key: Dict[str, str] = {}
  def _alias(src: str, target_label: str) -> None:
    norm_src = _normalize(src)
    if target_label in [lbl for lbl, _ in sum(SECTION_FIELDS["AKTIF"].values(), []) + sum(SECTION_FIELDS["PASIF"].values(), [])]:
      key = label_to_key.get(_normalize(target_label))
      if key:
        alias_to_key[norm_src] = key

  # Variations
  _alias("Donem Kari Vergi Yukumlulugu", "Dönem Karı Vergi Yükümlülüğü")
  _alias("Donem Kar Vergi Yukumlulugu", "Dönem Karı Vergi Yükümlülüğü")
  _alias("Cari Donem Vergisiyle ilgili Varliklar", "Cari Dönem Vergisiyle İlgili Varlıklar")
  _alias("Cari Donem Vergisiyle ilgili Borclar", "Cari Dönem Vergisiyle İlgili Borçlar")
  _alias("Ozkaynak Yontemiyle Degerlenen Yatirimlardan Yukumlulukler", "Özkaynak Yöntemiyle Değerlenen Yatırımlardan Yükümlülükler")
  _alias("Ozkaynak Yontemiyle Degerlenen Yatirimlar", "Özkaynak Yöntemiyle Değerlenen Yatırımlar")
  _alias("Pesin Odenmis Giderler", "Peşin Ödenmiş Giderler")
  _alias("Ertelenmis Vergi Varligi", "Ertelenmiş Vergi Varlığı")
  _alias("Ertelenmis Gelirler", "Ertelenmiş Gelirler")
  # TCMB label variants seen in files
  _alias("Turkiye Cumhuriyeti Merkez Bankasi Hesabi", "Türkiye Cumhuriyet Merkez Bankası Hesabı")
  _alias("Türkiye Cumhuriyeti Merkez Bankası Hesabı", "Türkiye Cumhuriyet Merkez Bankası Hesabı")

  def _pick_label_value(row_vals: Tuple) -> Tuple[str | None, float | None]:
    # Try to infer label and numeric value from an arbitrary row
    if not row_vals:
      return (None, None)
    non_empty = [(i, v) for i, v in enumerate(row_vals, start=1) if v is not None and str(v).strip() != ""]
    if not non_empty:
      return (None, None)
    # Prefer last numeric-looking value as amount
    value: float | None = None
    for _, v in reversed(non_empty):
      try:
        value = float(v)
        break
      except Exception:
        continue
    # Pick first string-like cell as label
    label: str | None = None
    for _, v in non_empty:
      if isinstance(v, str) and v.strip() != "":
        label = str(v)
        break
    return (label, value)

  # Keep track of the current detected group/section while scanning rows
  current_group: str | None = None

  for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
    # guard against short tuples
    if not row:
      continue
    
    if is_simple_format:
      # Simple format: just label and value
      label = row[0] if len(row) >= 1 else None
      val = row[1] if len(row) >= 2 else None
      if label is None:
        continue
      # Map by label only
      norm = _normalize(str(label))
      # Detect and update current group headers
      if norm in group_names_norm and (val is None or str(val).strip() == ""):
        current_group = norm
        continue
      # Resolve with group preference if available
      mapped = None
      if current_group and norm in group_label_to_key_map.get(current_group, {}):
        mapped = group_label_to_key_map[current_group][norm]
      else:
        mapped = label_to_key.get(norm)
      if mapped is None and not STRICT_LABELS:
        # İsteğe bağlı: fuzzy eşleştirme (kapalı)
        choices = list(label_to_key.keys())
        matches = difflib.get_close_matches(norm, choices, n=1, cutoff=0.75)
        if matches:
          mapped = label_to_key[matches[0]]
      if mapped is not None:
        try:
          d[str(mapped)] = float(val) if val is not None and str(val).strip() != "" else 0.0
        except Exception:
          d[str(mapped)] = 0.0
    else:
      # Detailed/unknown format: try configured columns else infer
      key = row[key_col - 1] if len(row) >= key_col else None
      val = row[val_col - 1] if len(row) >= val_col else None
      label = row[label_col - 1] if len(row) >= label_col else None
      # Update current group if a header row is encountered
      if label is not None:
        norm_label_only = _normalize(str(label))
        if norm_label_only in group_names_norm and (val is None or str(val).strip() == "") and (key is None or str(key).strip() == ""):
          current_group = norm_label_only
          continue
      if key is None and (label is None or (val is None or str(val).strip() == "")):
        # Infer from row dynamically
        label, val = _pick_label_value(row)
      if key is None:
        # fallback to label mapping (exact normalized, then fuzzy)
        if label is not None:
          norm = _normalize(str(label))
          mapped = None
          if current_group and norm in group_label_to_key_map.get(current_group, {}):
            mapped = group_label_to_key_map[current_group][norm]
          else:
            mapped = label_to_key.get(norm)
          if mapped is None:
            mapped = alias_to_key.get(norm)
          if mapped is None and not STRICT_LABELS:
            # İsteğe bağlı: fuzzy eşleştirme (kapalı)
            choices = list(label_to_key.keys())
            matches = difflib.get_close_matches(norm, choices, n=1, cutoff=0.82)
            if matches:
              mapped = label_to_key[matches[0]]
          key = mapped
      if key is None:
        continue
      try:
        d[str(key)] = float(val) if val is not None and str(val).strip() != "" else 0.0
      except Exception:
        d[str(key)] = 0.0
  return d


SECTION_FIELDS: Dict[str, Dict[str, List[Tuple[str, str]]]] = {
  "AKTIF": {
    "Dönen Varlıklar": [
      ("Nakit ve Nakit Benzerleri", "nakitVeNakitBenzerleri"),
      ("Gayrimenkul Projeleri Kapsamında Açılan Nakit Hesapları", "gayrimenkulProjeleriNakitHesaplari"),
      ("Finansal Yatırımlar", "finansalYatirimlar"),
      ("Teminata Verilen Finansal Varlıklar", "teminataVerilenFinansalVarliklar"),
      ("Ticari Alacaklar", "ticariAlacaklarDonen"),
      ("Finans Sektörü Faaliyetlerinden Alacaklar", "finansSektoruFaaliyetlerindenAlacaklarDonen"),
      ("Türkiye Cumhuriyet Merkez Bankası Hesabı", "tcmbHesabi"),
      ("Diğer Alacaklar", "digerAlacaklarDonen"),
      ("Müşteri Sözleşmelerinden Doğan Varlıklar", "musteriSozlesmelerindenDoganVarliklarDonen"),
      ("İmtiyaz Sözleşmelerine İlişkin Finansal Varlıklar", "imtiyazSozlesmelerineIliskinFinansalVarliklarDonen"),
      ("Türev Araçlar", "turevAraclarDonen"),
      ("Stoklar", "stoklar"),
      ("Proje Halindeki Stoklar", "projeHalindekiStoklar"),
      ("Canlı Varlıklar", "canliVarliklarDonen"),
      ("Peşin Ödenmiş Giderler", "pesinOdenmisGiderlerDonen"),
      ("Ertelenmiş Sigortacılık Üretim Giderleri", "ertelenmisSigortacilikUretimGiderleri"),
      ("Cari Dönem Vergisiyle İlgili Varlıklar", "cariDonemVergisiyleIlgiliVarliklarDonen"),
      ("Nakdi Dışı Serbest Kullanılabilir Teminatlar", "nakdiDisiSerbestKullanilabilirTeminatlarDonen"),
      ("Diğer Dönen Varlıklar", "digerDonenVarliklar"),
      ("Satış Amacıyla Elde Tutulan Duran Varlıklar", "satisAmaciylaEldeTutulanDuranVarliklar"),
      ("Ortaklara Dağıtılmak Üzere Elde Tutulan Duran Varlıklar", "ortaklaraDagitilmakUzereEldeTutulanDuranVarliklar"),
      ("Toplam Dönen Varlıklar", "toplamDonenVarliklar_dummy"),
    ],
    "Duran Varlıklar": [
      ("Finansal Yatırımlar", "finansalYatirimlarDuran"),
      ("İştirakler, İş Ortaklıkları ve Bağlı Ortaklıklardaki Yatırımlar", "istirakIsOrtaklikBagliOrtaklikYatirimlari"),
      ("Ticari Alacaklar", "ticariAlacaklar"),
      ("Finans Sektörü Faaliyetlerinden Alacaklar", "finansSektoruFaaliyetlerindenAlacaklarDuran"),
      ("Diğer Alacaklar", "digerAlacaklarDuran"),
      ("Müşteri Sözleşmelerinden Doğan Varlıklar", "musteriSozlesmelerindenDoganVarliklarDuran"),
      ("İmtiyaz Sözleşmelerine İlişkin Finansal Varlıklar", "imtiyazSozlesmelerineIliskinFinansalVarliklarDuran"),
      ("Türev Araçlar", "turevAraclarDuran"),
      ("Stoklar", "stoklarDuran"),
      ("Özkaynak Yöntemiyle Değerlenen Yatırımlar", "ozkaynakYontemiyleDegerlenenYatirimlar"),
      ("Canlı Varlıklar", "canliVarliklarDuran"),
      ("Yatırım Amaçlı Gayrimenkuller", "yatirimAmacliGayrimenkuller"),
      ("Proje Halindeki Yatırım Amaçlı Gayrimenkuller", "projeHalindekiYatirimAmacliGayrimenkuller"),
      ("Maddi Duran Varlıklar", "maddiDuranVarliklar"),
      ("Kullanım Hakkı Varlıkları", "kullanimHakkiVarliklari"),
      ("Maddi Olmayan Duran Varlıklar", "maddiOlmayanDuranVarliklar"),
      ("Peşin Ödenmiş Giderler", "pesinOdenmisGiderlerDuran"),
      ("Ertelenmiş Vergi Varlığı", "ertelenmisVergiVarligiDuran"),
      ("Cari Dönem Vergisiyle İlgili Duran Varlıklar", "cariDonemVergisiyleIlgiliDuranVarliklar"),
      ("Nakdi Dışı Serbest Kullanılabilir Teminatlar", "nakdiDisiSerbestKullanilabilirTeminatlarDuran"),
      ("Diğer Duran Varlıklar", "digerDuranVarliklar"),
      ("Toplam Duran Varlıklar", "toplamDuranVarliklar_dummy"),
    ],
  },
  "PASIF": {
    "Kısa Vadeli Yükümlülükler": [
      ("Finansal Borçlar", "finansalBorclarKV"),
      ("Diğer Finansal Yükümlülükler", "digerFinansalYukumluluklerKV"),
      ("Ticari Borçlar", "ticariBorclarKV"),
      ("Finans Sektörü Faaliyetlerinden Borçlar", "finansSektoruFaaliyetlerindenBorclarKV"),
      ("Çalışanlara Sağlanan Faydalar Kapsamında Borçlar", "calisanlaraSaglananFaydalarBorclarKV"),
      ("Diğer Borçlar", "digerBorclarKV"),
      ("Müşteri Sözleşmelerinden Doğan Yükümlülükler", "musteriSozlesmelerindenDoganYukumluluklerKV"),
      ("Özkaynak Yöntemiyle Değerlenen Yatırımlardan Yükümlülükler", "ozkaynakYontemiyleDegerlenenYatirimlardanYukumluluklerKV"),
      ("Türev Araçlar", "turevAraclarKV"),
      ("Devlet Teşvik ve Yardımları", "devletTesvikYardimKV"),
      ("Ertelenmiş Gelirler", "ertelenmisGelirlerKV"),
      ("Dönem Karı Vergi Yükümlülüğü", "donemKariVergiYukumluluguKV"),
      ("Kısa Vadeli Karşılıklar", "kisaVadeliKarsiliklar"),
      ("Diğer Kısa Vadeli Yükümlülükler", "digerKisaVadeliYukumlulukler"),
      ("Satış Amaçlı Sınıflandırılan Varlık Gruplarına İlişkin Yükümlülükler", "satisAmacliSiniflandirilanVarlikGruplarinaIliskinYukumlulukler"),
      ("Ortaklara Dağıtılmak Üzere Elde Tutulan Varlık Gruplarına İlişkin Yükümlülükler", "ortaklaraDagitilmakUzereEldeTutulanVarlikGruplarinaIliskinYukumlulukler"),
      ("Toplam Kısa Vadeli Yükümlülükler", "toplamKVY_dummy"),
    ],
    "Uzun Vadeli Yükümlülükler": [
      ("Finansal Borçlar", "finansalBorclarUV"),
      ("Diğer Finansal Yükümlülükler", "digerFinansalYukumluluklerUV"),
      ("Ticari Borçlar", "ticariBorclarUV"),
      ("Finans Sektörü Faaliyetlerinden Borçlar", "finansSektoruFaaliyetlerindenBorclarUV"),
      ("Çalışanlara Sağlanan Faydalar Kapsamında Borçlar", "calisanlaraSaglananFaydalarBorclarUV"),
      ("Diğer Borçlar", "digerBorclarUV"),
      ("Müşteri Sözleşmelerinden Doğan Yükümlülükler", "musteriSozlesmelerindenDoganYukumluluklerUV"),
      ("Özkaynak Yöntemiyle Değerlenen Yatırımlardan Yükümlülükler", "ozkaynakYontemiyleDegerlenenYatirimlardanYukumluluklerUV"),
      ("Türev Araçlar", "turevAraclarUV"),
      ("Devlet Teşvik ve Yardımları", "devletTesvikYardimUV"),
      ("Ertelenmiş Gelirler", "ertelenmisGelirlerUV"),
      ("Cari Dönem Vergisiyle İlgili Borçlar", "cariDonemVergisiyleIlgiliBorclarUV"),
      ("Ertelenmiş Vergi Yükümlülüğü", "ertelenmisVergiYukumlulugu"),
      ("Diğer Uzun Vadeli Yükümlülükler", "digerUzunVadeliYukumlulukler"),
      ("Toplam Uzun Vadeli Yükümlülükler", "toplamUVY_dummy"),
    ],
    "Özkaynaklar": [
      ("Ana Ortaklığa Ait Özkaynaklar", "anaOrtakligaAitOzkaynaklar"),
      ("Ödenmiş Sermaye", "odenmisSermaye"),
      ("Sermaye Düzeltme Farkları", "sermayeDuzeltmeFarklari"),
      ("Birleşme Dengeleştirme Hesabı", "birlesmeDengeletirmeHesabi"),
      ("Pay Sahiplerinin İlave Sermaye Katkıları", "paySahipleriIlaveSermayeKatkilari"),
      ("Sermaye Avansı", "sermayeAvansi"),
      ("Geri Alınmış Paylar (-)", "geriAlinmisPaylar"),
      ("Karşılıklı İştirak Sermaye Düzeltmesi (-)", "karsilikliIstirakSermayeDuzeltmesi"),
      ("Paylara İlişkin Primler (iskontolar)", "paylaraIliskinPrimlerIskontolar"),
      ("Ortak Kontrole Tabi Teşebbüs veya İşletmeleri İçeren Birleşmelerin Etkisi", "ortakKontroleTabiBirlesmeEtkisi"),
      ("Pay Bazlı Ödemeler (-)", "payBazliOdemeler"),
      ("Kar veya Zararda Yeniden Sınıflandırılmayacak Birikmiş Diğer Kapsamlı Gelirler (Giderler)", "yenidenSiniflandirilmayacakBirikmisDigerKapsamliGelirGider"),
      ("Kar veya Zararda Yeniden Sınıflandırılacak Birikmiş Diğer Kapsamlı Gelirler (Giderler)", "yenidenSiniflandirilacakBirikmisDigerKapsamliGelirGider"),
      ("Kardan Ayrılan Kısıtlanmış Yedekler", "kardanAyrilanKisitlanmisYedekler"),
      ("Diğer Yedekler", "digerYedekler"),
      ("Geçmiş Yıllar Kar/Zararları", "gecmisYillarKarZarari"),
      ("Dönem Net Karı", "donemKariZarari"),
      ("Azınlık Payları", "kontrolGucuOlmayanPaylar"),
      ("Toplam Özkaynaklar", "ozkaynaklarToplam_dummy"),
      ("Toplam Kaynaklar", "toplamKaynaklar_dummy"),
      ("Hedge Dahil Net Yabancı Para Pozisyonu", "hedgeDahilNetYabanciParaPozisyonu"),
    ],
  },
}

# Section key lists for totals
DONEN_KEYS = [k for _, k in SECTION_FIELDS["AKTIF"]["Dönen Varlıklar"] if not k.endswith("_dummy")]
DURAN_KEYS = [k for _, k in SECTION_FIELDS["AKTIF"]["Duran Varlıklar"] if not k.endswith("_dummy")]
KV_KEYS = [k for _, k in SECTION_FIELDS["PASIF"]["Kısa Vadeli Yükümlülükler"] if not k.endswith("_dummy")]
UV_KEYS = [k for _, k in SECTION_FIELDS["PASIF"]["Uzun Vadeli Yükümlülükler"] if not k.endswith("_dummy")]
OZ_KEYS = [k for _, k in SECTION_FIELDS["PASIF"]["Özkaynaklar"] if not k.endswith("_dummy")]


def gui_collect_data(vars_map: Dict[str, tk.Variable]) -> Dict[str, float | str]:
  d: Dict[str, float | str] = {}
  for key, var in vars_map.items():
    if key in ("isletmeAdi", "bilancoTarihi"):
      d[key] = str(var.get())
    else:
      d[key] = parse_float(str(var.get()))
  return d


def draw_canvas(canvas: tk.Canvas, aktif: float, pasif: float) -> None:
  canvas.delete("all")
  width = int(canvas.cget("width"))
  height = int(canvas.cget("height"))

  margin = 20
  bar_height = 40
  gap = 30

  max_val = max(aktif, pasif, 1)
  scale = (width - margin * 2) / max_val

  # Aktif barı
  aktif_len = aktif * scale
  canvas.create_rectangle(margin, margin, margin + aktif_len, margin + bar_height, fill="#16a34a", outline="")
  canvas.create_text(margin, margin - 5, anchor="sw", text=f"Aktif: {format_tl(aktif)} TL", fill="#065f46", font=("Segoe UI", 10, "bold"))

  # Pasif barı
  y2 = margin + bar_height + gap
  pasif_len = pasif * scale
  canvas.create_rectangle(margin, y2, margin + pasif_len, y2 + bar_height, fill="#2563eb", outline="")
  canvas.create_text(margin, y2 - 5, anchor="sw", text=f"Pasif: {format_tl(pasif)} TL", fill="#1e3a8a", font=("Segoe UI", 10, "bold"))

  # Denge göstergesi
  diff = aktif - pasif
  status = "DENGELİ" if abs(diff) <= 0.01 else ("AKTİF > PASİF" if diff > 0 else "PASİF > AKTİF")
  color = "#15803d" if status == "DENGELİ" else ("#b45309" if diff > 0 else "#b91c1c")
  canvas.create_text(width - margin, height - margin, anchor="se", text=f"Durum: {status} | Fark: {format_tl(diff)} TL", fill=color, font=("Segoe UI", 10, "bold"))


def create_gui() -> int:
  if tk is None:
    print("Tkinter yüklenemedi; GUI kullanılamıyor. CLI'ı çalıştırın veya Python Tk desteğini kurun.")
    return 2

  root = tk.Tk()
  root.title("Bilanço Tablosu - Masaüstü Arayüz")
  # Tam ekran/maximum başlat
  try:
    root.state("zoomed")  # Windows'ta maksimum boyut
  except Exception:
    try:
      root.attributes("-zoomed", True)  # Bazı platformlarda
    except Exception:
      root.attributes("-fullscreen", True)
  root.minsize(1024, 700)

  # F11 ile fullscreen, Esc ile çıkış
  def _toggle_fullscreen(event=None):
    is_full = bool(root.attributes("-fullscreen"))
    root.attributes("-fullscreen", not is_full)
  def _end_fullscreen(event=None):
    root.attributes("-fullscreen", False)
  root.bind("<F11>", _toggle_fullscreen)
  root.bind("<Escape>", _end_fullscreen)

  container = ttk.Frame(root, padding=8)
  container.grid(row=0, column=0, sticky="nsew")
  root.rowconfigure(0, weight=1)
  root.columnconfigure(0, weight=1)

  # Üst başlık (işletme adı ve tarih)
  header = ttk.Frame(container)
  header.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
  ttk.Label(header, text="İşletme Adı:").grid(row=0, column=0, sticky="w")
  isletme_var = tk.StringVar(value="")
  ttk.Entry(header, textvariable=isletme_var, width=40).grid(row=0, column=1, sticky="w", padx=(4, 16))
  ttk.Label(header, text="Tarih (YYYY-MM-DD):").grid(row=0, column=2, sticky="w")
  tarih_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
  ttk.Entry(header, textvariable=tarih_var, width=16).grid(row=0, column=3, sticky="w", padx=4)

  # Sağ üst LOG paneli
  log_frame = ttk.LabelFrame(header, text="Log")
  log_frame.grid(row=0, column=4, sticky="ne", padx=(12, 0))
  log_text = tk.Text(log_frame, width=52, height=6, state="disabled", wrap="word")
  log_text.grid(row=0, column=0)
  log_text.tag_configure("info", foreground="#1f2937")
  log_text.tag_configure("ok", foreground="#166534")
  log_text.tag_configure("warn", foreground="#b45309")
  log_text.tag_configure("err", foreground="#b91c1c")

  def log(message: str, kind: str = "info") -> None:
    timestamp = datetime.now().strftime("%H:%M:%S")
    try:
      log_text.configure(state="normal")
      log_text.insert("end", f"[{timestamp}] {message}\n", kind if kind in ("info", "ok", "warn", "err") else "info")
      log_text.see("end")
    finally:
      log_text.configure(state="disabled")

  # Scroll alanı ve iki sütunlu düzen
  # Canvas + iç frame ile kaydırma
  scroll_container = ttk.Frame(container)
  scroll_container.grid(row=1, column=0, columnspan=2, sticky="nsew")
  container.rowconfigure(1, weight=1)
  container.columnconfigure(0, weight=1)
  container.columnconfigure(1, weight=1)

  canvas = tk.Canvas(scroll_container, borderwidth=0, highlightthickness=0)
  vsb = ttk.Scrollbar(scroll_container, orient="vertical", command=canvas.yview)
  inner = ttk.Frame(canvas)
  inner_id = canvas.create_window((0, 0), window=inner, anchor="nw")
  canvas.configure(yscrollcommand=vsb.set)
  canvas.grid(row=0, column=0, sticky="nsew")
  vsb.grid(row=0, column=1, sticky="ns")
  scroll_container.rowconfigure(0, weight=1)
  scroll_container.columnconfigure(0, weight=1)

  def _on_configure(event):
    canvas.configure(scrollregion=canvas.bbox("all"))
    canvas.itemconfigure(inner_id, width=event.width - vsb.winfo_width())
  canvas.bind("<Configure>", _on_configure)

  # Hücre oluşturucu (çizgili tablo görünümü için)
  def cell(parent, text="", bold=False, entry_var: tk.StringVar | None = None):
    style = {"borderwidth":1, "relief":"solid"}
    if entry_var is None:
      lbl = tk.Label(parent, text=text, anchor="w", padx=6, pady=4, **style)
      if bold:
        lbl.configure(font=("Segoe UI", 10, "bold"))
      return lbl
    ent = tk.Entry(parent, textvariable=entry_var, justify="right", **style)
    return ent

  vars_map: Dict[str, tk.Variable] = {"isletmeAdi": isletme_var, "bilancoTarihi": tarih_var}
  for side in ("AKTIF", "PASIF"):
    for group, fields in SECTION_FIELDS[side].items():
      for label, key in fields:
        if key.endswith("_dummy"):
          continue
        if key not in vars_map:
          vars_map[key] = tk.StringVar(value="0")

  # İçerik: AKTİF ve PASİF sütunları
  col_left = ttk.Frame(inner)
  col_right = ttk.Frame(inner)
  col_left.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
  col_right.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
  inner.columnconfigure(0, weight=1)
  inner.columnconfigure(1, weight=1)

  def render_group(parent, title: str, fields: List[Tuple[str, str]]):
    box = ttk.LabelFrame(parent, text=title)
    box.grid(sticky="nsew", pady=(0, 6))
    r = 0
    for label, key in fields:
      if key.endswith("_dummy"):
        continue
      cell(box, label).grid(row=r, column=0, sticky="nsew")
      cell(box, entry_var=vars_map[key]).grid(row=r, column=1, sticky="nsew")
      r += 1
    box.columnconfigure(1, weight=1)

  # AKTİF
  ttk.Label(col_left, text="AKTİF", font=("Segoe UI", 10, "bold")).grid(sticky="w", pady=(0, 4))
  for group, fields in SECTION_FIELDS["AKTIF"].items():
    render_group(col_left, group, fields)

  # PASİF
  ttk.Label(col_right, text="PASİF", font=("Segoe UI", 10, "bold")).grid(sticky="w", pady=(0, 4))
  for group, fields in SECTION_FIELDS["PASIF"].items():
    render_group(col_right, group, fields)

  # Alt toplam satırları
  aktif_top_lbl = tk.Label(container, text="Aktif (Varlıklar) Toplamı", anchor="w", padx=6, pady=6, borderwidth=1, relief="solid", font=("Segoe UI", 10, "bold"))
  aktif_top_val = tk.Label(container, text="0,00", anchor="e", padx=6, pady=6, borderwidth=1, relief="solid", font=("Segoe UI", 10, "bold"))
  pasif_top_lbl = tk.Label(container, text="Pasif (Kaynaklar) Toplamı", anchor="w", padx=6, pady=6, borderwidth=1, relief="solid", font=("Segoe UI", 10, "bold"))
  pasif_top_val = tk.Label(container, text="0,00", anchor="e", padx=6, pady=6, borderwidth=1, relief="solid", font=("Segoe UI", 10, "bold"))

  aktif_top_lbl.grid(row=2, column=0, sticky="nsew", pady=(8,0))
  aktif_top_val.grid(row=2, column=1, sticky="nsew", pady=(8,0))
  pasif_top_lbl.grid(row=3, column=0, sticky="nsew")
  pasif_top_val.grid(row=3, column=1, sticky="nsew")

  # Alt butonlar (Doğrula / JSON Kaydet / MongoDB Kaydet)
  buttons = ttk.Frame(container)
  buttons.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(8,0))

  def build_dict() -> Dict[str, float | str]:
    return gui_collect_data(vars_map)

  def do_open_json() -> None:
    if filedialog is None:
      messagebox.showerror("Dosya", "Dosya iletişim kutusu kullanılamıyor.")
      return
    path = filedialog.askopenfilename(
      title="Bilanço JSON Aç",
      filetypes=(("JSON Files", "*.json"), ("All Files", "*.*"))
    )
    if not path:
      return
    try:
      with open(path, "r", encoding="utf-8") as f:
        doc = json.load(f)
    except Exception as ex:
      messagebox.showerror("Dosya", f"JSON okunamadı: {ex}")
      log(f"JSON açma hatası: {ex}", "err")
      return

    # Temel bilgiler
    try:
      info = doc.get("isletmeBilgileri", {})
      if "ad" in info and "isletmeAdi" in vars_map:
        vars_map["isletmeAdi"].set(str(info.get("ad", "")))
      if "tarih" in info and "bilancoTarihi" in vars_map:
        vars_map["bilancoTarihi"].set(str(info.get("tarih", "")))
    except Exception:
      pass

    # Aktif
    try:
      akt = doc.get("aktif", {})
      donen = akt.get("donenVarliklar", {})
      duran = akt.get("duranVarliklar", {})
      for key in set(DONEN_KEYS + DURAN_KEYS + ["birikmiAmort"]):
        if key in donen:
          if key in vars_map:
            vars_map[key].set(str(donen.get(key)))
        if key in duran:
          if key in vars_map:
            vars_map[key].set(str(duran.get(key)))
    except Exception:
      pass

    # Pasif
    try:
      pas = doc.get("pasif", {})
      kv = pas.get("kisaVadeliYabanciKaynaklar", {})
      uv = pas.get("uzunVadeliYabanciKaynaklar", {})
      oz = pas.get("ozKaynaklar", {})
      for key in KV_KEYS:
        if key in kv and key in vars_map:
          vars_map[key].set(str(kv.get(key)))
      for key in UV_KEYS:
        if key in uv and key in vars_map:
          vars_map[key].set(str(uv.get(key)))
      for key in OZ_KEYS:
        if key in oz and key in vars_map:
          vars_map[key].set(str(oz.get(key)))
    except Exception:
      pass

    refresh()
    log(f"JSON yüklendi: {path}", "ok")

  def do_save_excel() -> None:
    d = build_dict()
    errs = validate(d)
    kritik = [e for e in errs if e.get("tip") == "kritik"]
    if kritik:
      text = "\n".join([f"- {e['tip'].upper()}: {e['mesaj']}" for e in errs])
      messagebox.showerror("Kritik Hatalar", text + "\n\nLütfen düzeltip tekrar deneyin.")
      log("Excel kaydedilmedi: kritik hatalar var.", "err")
      return
    if filedialog is None:
      messagebox.showerror("Dosya", "Dosya iletişim kutusu kullanılamıyor.")
      return
    path = filedialog.asksaveasfilename(
      title="Excel Olarak Kaydet",
      defaultextension=".xlsx",
      filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
    )
    if not path:
      return
    try:
      save_to_excel(path, d)
      messagebox.showinfo("Excel", f"Excel dosyası kaydedildi:\n{path}")
      log(f"Excel kaydedildi: {path}", "ok")
    except Exception as ex:
      messagebox.showerror("Excel", f"Kaydetme hatası: {ex}")
      log(f"Excel kaydı hatası: {ex}", "err")

  def do_open_excel() -> None:
    if filedialog is None:
      messagebox.showerror("Dosya", "Dosya iletişim kutusu kullanılamıyor.")
      return
    path = filedialog.askopenfilename(
      title="Excel Aç",
      filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
    )
    if not path:
      return
    try:
      data = load_from_excel(path)
    except Exception as ex:
      messagebox.showerror("Excel", f"Okuma hatası: {ex}")
      log(f"Excel açma hatası: {ex}", "err")
      return
    # fill vars
    updated = 0
    if "isletmeAdi" in data and "isletmeAdi" in vars_map:
      vars_map["isletmeAdi"].set(str(data.get("isletmeAdi", "")))
    if "bilancoTarihi" in data and "bilancoTarihi" in vars_map:
      vars_map["bilancoTarihi"].set(str(data.get("bilancoTarihi", "")))
    for key, var in vars_map.items():
      if key in ("isletmeAdi", "bilancoTarihi"):
        continue
      if key in data:
        var.set(str(data[key]))
        updated += 1
    refresh()
    log(f"Excel yüklendi: {path} (güncellenen alan: {updated})", "ok")

  def do_validate() -> None:
    errs = validate(build_dict())
    if not errs:
      messagebox.showinfo("Doğrulama", "Doğrulama başarılı, hata bulunamadı.")
      log("Doğrulama başarılı.", "ok")
    else:
      text = "\n".join([f"- {e['tip'].upper()}: {e['mesaj']}" for e in errs])
      messagebox.showwarning("Doğrulama Sonuçları", text)
      level = "err" if any(e.get("tip") == "kritik" for e in errs) else "warn"
      log("Doğrulama uyarıları/hataları var.", level)

  def do_save_json() -> None:
    d = build_dict()
    errs = validate(d)
    kritik = [e for e in errs if e.get("tip") == "kritik"]
    if kritik:
      text = "\n".join([f"- {e['tip'].upper()}: {e['mesaj']}" for e in errs])
      messagebox.showerror("Kritik Hatalar", text + "\n\nLütfen düzeltip tekrar deneyin.")
      log("JSON kaydedilmedi: kritik hatalar var.", "err")
      return
    doc = build_mongo_like_document(d, errs)
    filename = f"bilanco_tablolari_{int(datetime.now().timestamp())}.json"
    try:
      with open(filename, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)
      messagebox.showinfo("Kayıt Başarılı", f"JSON dosyası yazıldı:\n{filename}")
      log(f"JSON kaydedildi: {filename}", "ok")
    except Exception as ex:
      messagebox.showerror("Kayıt Hatası", f"Dosya yazılamadı: {ex}")
      log(f"JSON kaydı hatası: {ex}", "err")

  def do_save_mongo() -> None:
    d = build_dict()
    errs = validate(d)
    kritik = [e for e in errs if e.get("tip") == "kritik"]
    if kritik:
      text = "\n".join([f"- {e['tip'].upper()}: {e['mesaj']}" for e in errs])
      messagebox.showerror("Kritik Hatalar", text + "\n\nLütfen düzeltip tekrar deneyin.")
      log("MongoDB kaydedilmedi: kritik hatalar var.", "err")
      return
    doc = build_mongo_like_document(d, errs)
    uri = os.getenv("MONGO_URI", "mongodb://localhost:27017")
    dbn = os.getenv("MONGO_DB", "bilanco")
    coll = os.getenv("MONGO_COLLECTION", "bilanco_tablolari")
    try:
      inserted_id = save_to_mongo(doc, uri, dbn, coll)
      messagebox.showinfo("MongoDB Kayıt", f"MongoDB'ye kaydedildi.\n_id: {inserted_id}\n{dbn}.{coll}")
      log(f"MongoDB kaydedildi: {dbn}.{coll} _id={inserted_id}", "ok")
    except Exception as ex:
      messagebox.showerror("MongoDB Hatası", str(ex))
      log(f"MongoDB hatası: {ex}", "err")

  ttk.Button(buttons, text="Aç (JSON)", command=do_open_json).grid(row=0, column=0, padx=(0,6))
  ttk.Button(buttons, text="Aç (Excel)", command=do_open_excel).grid(row=0, column=1, padx=(0,6))
  
  def do_open_mongo() -> None:
    if MongoClient is None:
      messagebox.showerror("MongoDB", "PyMongo yüklü değil. 'pip install pymongo' ile yükleyin.")
      return
    uri = os.getenv("MONGO_URI", "mongodb://localhost:27017")
    dbn = os.getenv("MONGO_DB", "bilanco")
    coll = os.getenv("MONGO_COLLECTION", "bilanco_tablolari")
    client = MongoClient(uri)
    try:
      use_latest = messagebox.askyesno(
        "MongoDB Aç",
        "Son kaydı yüklemek ister misiniz?\nHayır derseniz belirli bir _id girebilirsiniz."
      )
      query_doc = None
      if use_latest:
        query_doc = client[dbn][coll].find_one(sort=[("kayitTarihi", -1)])
      else:
        if simpledialog is None:
          messagebox.showerror("MongoDB Aç", "_id girmek için gerekli diyalog açılamadı.")
          return
        entered = simpledialog.askstring("MongoDB Aç", "_id girin (boş bırakılırsa son kayıt yüklenir):")
        if entered and entered.strip() != "":
          query_doc = client[dbn][coll].find_one({"_id": entered.strip()})
        else:
          query_doc = client[dbn][coll].find_one(sort=[("kayitTarihi", -1)])
      if not query_doc:
        messagebox.showwarning("MongoDB Aç", "Belge bulunamadı.")
        return
      # Populate fields from document structure produced by build_mongo_like_document
      try:
        info = query_doc.get("isletmeBilgileri", {})
        if "ad" in info and "isletmeAdi" in vars_map:
          vars_map["isletmeAdi"].set(str(info.get("ad", "")))
        if "tarih" in info and "bilancoTarihi" in vars_map:
          vars_map["bilancoTarihi"].set(str(info.get("tarih", "")))
      except Exception:
        pass
      try:
        akt = query_doc.get("aktif", {})
        donen = akt.get("donenVarliklar", {})
        duran = akt.get("duranVarliklar", {})
        for key in set(DONEN_KEYS + DURAN_KEYS + ["birikmiAmort"]):
          if key in donen and key in vars_map:
            vars_map[key].set(str(donen.get(key)))
          if key in duran and key in vars_map:
            vars_map[key].set(str(duran.get(key)))
      except Exception:
        pass
      try:
        pas = query_doc.get("pasif", {})
        kv = pas.get("kisaVadeliYabanciKaynaklar", {})
        uv = pas.get("uzunVadeliYabanciKaynaklar", {})
        oz = pas.get("ozKaynaklar", {})
        for key in KV_KEYS:
          if key in kv and key in vars_map:
            vars_map[key].set(str(kv.get(key)))
        for key in UV_KEYS:
          if key in uv and key in vars_map:
            vars_map[key].set(str(uv.get(key)))
        for key in OZ_KEYS:
          if key in oz and key in vars_map:
            vars_map[key].set(str(oz.get(key)))
      except Exception:
        pass
      refresh()
      log("MongoDB belgesi yüklendi.", "ok")
    except Exception as ex:
      messagebox.showerror("MongoDB Aç", str(ex))
      log(f"MongoDB açma hatası: {ex}", "err")
    finally:
      try:
        client.close()
      except Exception:
        pass

  ttk.Button(buttons, text="Aç (MongoDB)", command=do_open_mongo).grid(row=0, column=2, padx=(0,6))
  ttk.Button(buttons, text="Doğrula", command=do_validate).grid(row=0, column=6, padx=(0,6))
  ttk.Button(buttons, text="Kaydet (JSON)", command=do_save_json).grid(row=0, column=3, padx=(0,6))
  ttk.Button(buttons, text="Kaydet (Excel)", command=do_save_excel).grid(row=0, column=4, padx=(0,6))
  ttk.Button(buttons, text="Kaydet (MongoDB)", command=do_save_mongo).grid(row=0, column=5)

  # Hesaplama ve canlı güncelleme
  def refresh(_: str | None = None, __: str | None = None, ___: str | None = None):
    d = build_dict()
    d["isletmeAdi"] = isletme_var.get()
    d["bilancoTarihi"] = tarih_var.get()
    aktif = sum_donen_varliklar(d) + sum_duran_varliklar(d)
    pasif = sum_kv_yabanci_kaynaklar(d) + sum_uv_yabanci_kaynaklar(d) + sum_oz_kaynaklar(d)
    aktif_top_val.configure(text=f"{format_tl(aktif)} TL")
    pasif_top_val.configure(text=f"{format_tl(pasif)} TL")

    # Kenarlık renkleri ile dengenin görsel ipucu
    if abs(aktif - pasif) <= 0.01:
      aktif_top_val.configure(fg="#15803d")
      pasif_top_val.configure(fg="#15803d")
    else:
      aktif_top_val.configure(fg="#b91c1c")
      pasif_top_val.configure(fg="#b91c1c")

  for var in vars_map.values():
    if isinstance(var, tk.Variable):
      var.trace_add("write", refresh)

  refresh()
  root.mainloop()
  return 0


def cli_main(args) -> int:
  print("BİLANÇO TABLOSU - Python CLI\n")
  data = collect_bilanco_data()

  print("\nHesaplama ve doğrulama yapılıyor...\n")
  errors = validate(data)

  kritik_sayisi = len([e for e in errors if e.get("tip") == "kritik"])
  if errors:
    print("Doğrulama Sonuçları:")
    for e in errors:
      print(f"- {e['tip'].upper()}: {e['mesaj']}")
  else:
    print("- Doğrulama başarılı, hata bulunamadı.")

  if kritik_sayisi > 0:
    print("\nKritik hatalar var! Lütfen düzeltip tekrar deneyin.")
    return 1

  doc = build_mongo_like_document(data, errors)

  filename = f"bilanco_tablolari_{doc['_id']}.json"
  with open(filename, "w", encoding="utf-8") as f:
    json.dump(doc, f, ensure_ascii=False, indent=2)

  print("\nBilanço başarıyla kaydedildi!")
  print(f"Belge ID: {doc['_id']}")
  print(f"Çıktı Dosyası: {filename}")

  if getattr(args, "mongo", False):
    uri = args.mongo_uri or os.getenv("MONGO_URI", "mongodb://localhost:27017")
    dbn = args.mongo_db or os.getenv("MONGO_DB", "bilanco")
    coll = args.mongo_coll or os.getenv("MONGO_COLLECTION", "bilanco_tablolari")
    try:
      inserted_id = save_to_mongo(doc, uri, dbn, coll)
      print(f"MongoDB'ye kaydedildi: {dbn}.{coll} _id={inserted_id}")
    except Exception as ex:
      print(f"MongoDB kaydetme hatası: {ex}")

  donen = sum_donen_varliklar(data)
  duran = sum_duran_varliklar(data)
  aktif = donen + duran
  kv = sum_kv_yabanci_kaynaklar(data)
  uv = sum_uv_yabanci_kaynaklar(data)
  oz = sum_oz_kaynaklar(data)
  pasif = kv + uv + oz

  print(f"\nAktif Toplamı: {format_tl(aktif)} TL")
  print(f"Pasif Toplamı: {format_tl(pasif)} TL")

  return 0


def main() -> int:
  parser = argparse.ArgumentParser(description="Bilanço Uygulaması (CLI + GUI)")
  parser.add_argument("--cli", action="store_true", help="GUI yerine CLI çalıştır")
  parser.add_argument("--mongo", action="store_true", help="CLI modunda MongoDB'ye de kaydet")
  parser.add_argument("--mongo-uri", dest="mongo_uri", default=None, help="MongoDB bağlantı URI (varsayılan env MONGO_URI veya mongodb://localhost:27017)")
  parser.add_argument("--mongo-db", dest="mongo_db", default=None, help="Veritabanı adı (varsayılan env MONGO_DB veya bilanco)")
  parser.add_argument("--mongo-coll", dest="mongo_coll", default=None, help="Koleksiyon adı (varsayılan env MONGO_COLLECTION veya bilanco_tablolari)")
  args = parser.parse_args()
  if args.cli:
    return cli_main(args)
  return create_gui()


if __name__ == "__main__":
  try:
    sys.exit(main())
  except KeyboardInterrupt:
    print("\nİptal edildi.")
    sys.exit(130)